import json
from calciatore import Calciatore

f = open("dataset/fantacalcio.json","r")
data = json.load(f)
f.close()

import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
stagione_19_20= wb.create_sheet("2019-2020")
stagione_20_21= wb.create_sheet("2020-2021")
stagione_21_22= wb.create_sheet("2021-2022")
stagione_22_23= wb.create_sheet("2022-2023")
stagione_23_24= wb.create_sheet("2023-2024")