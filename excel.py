#importazione pacchetti una volta installati e importazione file python
from calciatore import Calciatore
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class Excel:
    def __init__(self, calciatori:list[Calciatore]):
        self.calciatori: list[Calciatore] = calciatori
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  #rimuove la prima pagina vuota dell'excel
    
    
    def calciatori_attivi(self, stagione)->list[Calciatore]:
        """aggiunge alla ista i calciatori solo se sono attivi in quella stagione"""
        lista_calciatori=[]
        for calciatore in self.calciatori:
            if calciatore.attivo(stagione):
                lista_calciatori.append(calciatore)
        return lista_calciatori
    
    def crea_foglio_excel(self, calciatori:list[Calciatore], anno:str):
        """Creaziione di un foglio excel in cui ci sono i calciatori attivi in un determinato anno(stagione)"""
        for calciatore in calciatori:
            if calciatore.attivo(anno):
                self.wb.create_sheet

    def scrivi_statistiche(self, stagioni:list[str]):
        """ inserisce i dati dei calciatori nei foglio <stagioni> excel se sono attivi nella stagione"""
        #modifica dell'header
        header_fill = PatternFill(start_color = "4F81BD", end_color = "4F81BD",  fill_type = "solid")
        header_font= Font(bold = True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
        green_fill=PatternFill(start_color="C6F5D6",end_color="C6F5D6", fill_type="solid" )

        for stagione in stagioni:
            sheet= self.wb[stagione] #prende il foglio dal nome stagione
            if sheet:
                headers= ["Nome", "Media_voti", "MediaFanta", "Squadra", "Ruolo"]
                
                #intestazioni:
                for col, header in enumerate(headers, start=1): 
                   cell = sheet.cell(row=1, column=col)
                   cell.value= header
                   cell.fill = header_fill
                   cell.font= header_font
                   cell.alignment = Alignment(horizontal="center")
                   cell.border = thin_border 
                                       
                sheet.freeze_panes = "A2" #blocca headers in alto

                #larghezza colonne
                sheet.column_dimensions["A"].width = 20  
                sheet.column_dimensions["B"].width = 15
                sheet.column_dimensions["C"].width = 15
                sheet.column_dimensions["D"].width = 15
                sheet.column_dimensions["E"].width = 10

                riga = 2
                for calciatore in self.calciatori:
                    if calciatore.attivo(stagione):
                         #evidenzia la riga di verde se il voto è maggiore o uguale a 6
                        if calciatore.media_voti[stagione] >= 6:
                                for col in range(1, 6):
                                    sheet.cell(row=riga, column=col).fill = green_fill

                        sheet.cell(row=riga, column=1).value = calciatore.nome
                        sheet.cell(row=riga, column=2).value = calciatore.media_voti[stagione]
                        sheet.cell(row=riga, column=3).value = calciatore.media_voti_fanta[stagione]
                        sheet.cell(row=riga, column=4).value = calciatore.squadra
                        sheet.cell(row=riga, column=5).value = calciatore.ruolo
                        riga += 1


    def crea_fogli_per_stagioni(self, stagioni:list[str]):    
        """Creazione fogli excel per ogni anno con i suoi calciatori attivi"""
        for stagione in stagioni:
            self.wb.create_sheet(stagione)  
        return

    def save_to_file(self,file_path:str):
        #salvataggio file excel
        self.wb.save(file_path)

    